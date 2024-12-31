from django.contrib.gis.geos import Point

import openpyxl

from results import models, results_util

def parse_simplemaps_us(path: str):
	wb_xls = openpyxl.load_workbook(path)
	if len(wb_xls.worksheets) > 1:
		raise RuntimeError(f'Strange number of sheets in {path}: {len(wb_xls.worksheets)}')
	sheet = wb_xls.worksheets[0]
	if sheet.max_column != 17:
		raise RuntimeError(f'Strange number of columns in {path}: want 17, got {sheet.max_column}')

	cities_added = 0
	regions_added = 0
	errors = []
	existing_cities_by_id = {city.simplemaps_id: city for city in models.City.objects.exclude(simplemaps_id=None)}
	ids_met = set()
	ids_existing = 0
	print(f'rows: {sheet.max_row}')
	for row in range(2, sheet.max_row + 1):
	# for row in range(2, 20):
		if (row % 1000) == 0:
			print(f'Processed rows until {row-1}. Met {len(ids_met)} ids, {ids_existing} already existed')
		simplemaps_id = results_util.int_safe(sheet.cell(row, 17).value)
		if not (1000 <= simplemaps_id <= 2147483647):
			print(f'id {simplemaps_id} in row {row} is too big')
		if simplemaps_id in ids_met:
			print(f'id {simplemaps_id} is in row {row} and was before')
		ids_met.add(simplemaps_id)
		name = sheet.cell(row, 1).value.strip()
		if simplemaps_id in existing_cities_by_id:
			city = existing_cities_by_id[simplemaps_id]
			if city.name != name:
				errors.append(f'row {row}: city with simplemaps_id {simplemaps_id} is {name} in file but {existing_cities_by_id[simplemaps_id]} in DB')
				print(errors[-1])
				if len(errors) > 5:
					break
			ids_existing += 1
			continue

		# print(f'row {row}: reached this place')
		state = sheet.cell(row, 4).value
		state_id = sheet.cell(row, 3).value
		if len(state_id) != 2:
			raise RuntimeError(f'Row {row}: strange state_id {state_id}')
		if state[0] != state_id[0]:
			raise RuntimeError(f'Row {row}: strange state_id {state_id}, state {state}')
		if state == 'Puerto Rico':
			region = models.Region.objects.get(country__name='Puerto Rico')
		else:
			country = models.Country.objects.get(pk='US')
			region, just_created = models.Region.objects.get_or_create(country=country, name=state)
			if just_created:
				regions_added += 1
				models.Region_conversion.objects.create(country=country, region=region, region_raw=state)
				models.Region_conversion.objects.create(country=country, region=region, country_raw=country.name, region_raw=state_id)
				models.Region_conversion.objects.create(country=country, region=region, country_raw=country.id, region_raw=state_id)

		county = sheet.cell(row, 6).value.strip()
		if not county:
			errors.append(f'row {row}: no county')
		county += ' County'
		city, just_created = models.City.objects.get_or_create(name=name, region=region, raion=county)
		fields_touched = []
		if just_created:
			cities_added += 1
			fields_touched = ['name', 'region']
		elif city.simplemaps_id and (city.simplemaps_id != simplemaps_id):
			errors.append(f'row {row}: simplemaps_id {simplemaps_id} is different from city id {city.id} has')
		if not city.geo:
			city.geo = Point(sheet.cell(row, 8).value, sheet.cell(row, 7).value)
			fields_touched.append('geo')
		if not city.population:
			city.population = sheet.cell(row, 9).value
			fields_touched.append('population')
		if not city.simplemaps_id:
			city.simplemaps_id = simplemaps_id
			fields_touched.append('simplemaps_id')
		if fields_touched:
			city.save()
			models.log_obj_create(models.USER_ROBOT_CONNECTOR, city, models.ACTION_CREATE if just_created else models.ACTION_UPDATE, field_list=fields_touched)
	print(f'Done! {cities_added} new cities, met {len(ids_met)} ids, {ids_existing} already existed, {regions_added} new regions, {len(errors)} errors')
	for i, error in enumerate(errors):
		print(f'{i+1}. {error}')

def parse_simplemaps_world(path: str):
	wb_xls = openpyxl.load_workbook(path)
	if len(wb_xls.worksheets) > 1:
		raise RuntimeError(f'Strange number of sheets in {path}: {len(wb_xls.worksheets)}')
	sheet = wb_xls.worksheets[0]
	if sheet.max_column != 11:
		raise RuntimeError(f'Strange number of columns in {path}: want 11, got {sheet.max_column}')

	cities_added = 0
	regions_added = 0
	errors = []
	existing_cities_by_id = {city.simplemaps_id: city for city in models.City.objects.exclude(simplemaps_id=None)}
	ids_met = set()
	ids_existing = 0
	print(f'rows: {sheet.max_row}')
	for row in range(2, sheet.max_row + 1):
	# for row in range(2, 4):
		if (row % 1000) == 0:
			print(f'Processed rows until {row-1}. Met {len(ids_met)} ids, {ids_existing} already existed')
		simplemaps_id = results_util.int_safe(sheet.cell(row, 11).value)
		if not (1000 <= simplemaps_id <= 2147483647):
			print(f'id {simplemaps_id} in row {row} is too big')
		if simplemaps_id in ids_met:
			print(f'id {simplemaps_id} is in row {row} and was before')
		ids_met.add(simplemaps_id)
		name = sheet.cell(row, 1).value.strip()
		if simplemaps_id in existing_cities_by_id:
			city = existing_cities_by_id[simplemaps_id]
			if city.name != name:
				errors.append(f'row {row}: city with simplemaps_id {simplemaps_id} is {name} in file but {existing_cities_by_id[simplemaps_id]} in DB')
				print(errors[-1])
				if len(errors) > 5:
					break
			ids_existing += 1
			continue

		# print(f'row {row}: reached this place')
		region_name = sheet.cell(row, 8).value or ''
		country_code = sheet.cell(row, 6).value or ''
		country = models.Country.objects.filter(pk=country_code).first() # iso2
		if not country:
			conversions = models.Country_conversion.objects.filter(country_raw=country_code)
			if conversions.count() == 1:
				country = conversions[0].country
			else:
				raise Exception(f'Row {row}: Country with ID "{sheet.cell(row, 6).value}" has {conversions.count()} possible conversions')
		# print(f'Row {row}: adding country_conversion "{sheet.cell(row, 7).value}" -> "{country.id}"')
		models.Country_conversion.objects.update_or_create(country=country, country_raw=sheet.cell(row, 7).value) # iso3
		region, just_created = models.Region.objects.get_or_create(country=country, name=region_name if region_name else country.name)
		if just_created:
			regions_added += 1
			models.Region_conversion.objects.update_or_create(country_raw=country.name, region_raw=region_name, country=country, region=region)

		if (count := models.City.objects.filter(name=name, region=region).count()) > 1:
			# Because e.g. https://en.wikipedia.org/wiki/Middletown,_Pennsylvania
			city, just_created = models.City.objects.get_or_create(name=name, region=region, simplemaps_id=simplemaps_id)
		else:
			city, just_created = models.City.objects.get_or_create(name=name, region=region)
		fields_touched = []
		if just_created:
			cities_added += 1
			fields_touched = ['name', 'region']
		elif city.simplemaps_id and (city.simplemaps_id != simplemaps_id):
			errors.append(f'row {row}: simplemaps_id {simplemaps_id} is different from what city {city.name} ({city.id}) has in DB: {city.simplemaps_id}')
		if not city.geo:
			city.geo = Point(sheet.cell(row, 4).value, sheet.cell(row, 3).value)
			fields_touched.append('geo')
		if not city.population:
			city.population = sheet.cell(row, 10).value
			fields_touched.append('population')
		if not city.simplemaps_id:
			city.simplemaps_id = simplemaps_id
			fields_touched.append('simplemaps_id')
		if fields_touched:
			city.save()
			models.log_obj_create(models.USER_ROBOT_CONNECTOR, city, models.ACTION_CREATE if just_created else models.ACTION_UPDATE, field_list=fields_touched)

		name_ascii = sheet.cell(row, 2).value
		if name_ascii and (name_ascii != name): # Bug: empty ascii for Dessau-Rosslau
			kwargs = {}
			if region.name != country.name:
				kwargs['region_raw'] = region.name
			models.City_conversion.objects.update_or_create(country_raw=country.name, city_raw=name_ascii, city=city, **kwargs)
	print(f'Done! {cities_added} new cities, met {len(ids_met)} ids, {ids_existing} already existed, {regions_added} new regions, {len(errors)} errors')
	for i, error in enumerate(errors):
		print(f'{i+1}. {error}')
